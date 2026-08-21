"""Puerta de entrada para XLSX de Empleabilidad con años variables."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from agente.normalizador.modelos import (
    Hallazgo,
    HojaInspeccion,
    ResultadoValidacionEntrada,
    TipoHoja,
)

ROLES: tuple[TipoHoja, ...] = ("convenios", "informes", "publicaciones")

COLUMNAS_OBLIGATORIAS: dict[TipoHoja, tuple[str, ...]] = {
    "convenios": (
        "RUC",
        "Empresa",
        "Facultad",
        "Cód_carrera",
        "Carrera",
        "Ciclo_convenio",
    ),
    "informes": (
        "Año",
        "Ciclo",
        "Facultad",
        "Cód_carrera",
        "Carrera",
    ),
    "publicaciones": (
        "Año de la vacante",
        "RUC",
        "Razón Social",
        "Identificación",
        "Tipo de puesto",
        "Fecha de publicación",
        "Fecha de finalización",
        "Creado_Empleo",
        "Área",
        "Área específica",
        "Cargo",
        "Cargo específico",
        "Posición a publicar",
        "Carrera resumen",
        "Funciones",
    ),
}


def normalizar_etiqueta(valor: object) -> str:
    """Normaliza una etiqueta para comparar nombres sin destruir la fuente."""

    texto = unicodedata.normalize("NFKD", str(valor or "")).lower()
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def _texto(valor: object) -> str:
    """Convierte una celda a texto limpio para inspección estructural."""

    return re.sub(r"\s+", " ", str(valor or "")).strip()


def detectar_rol(nombre_hoja: str) -> TipoHoja | None:
    """Identifica el universo por prefijo y permite cualquier rango de años."""

    etiqueta = normalizar_etiqueta(nombre_hoja)
    for rol in ROLES:
        if etiqueta == rol or etiqueta.startswith(f"{rol} "):
            return rol
    return None


def extraer_anios(nombre_hoja: str) -> tuple[int, ...]:
    """Extrae años declarados sin exigir un periodo concreto."""

    anios = {int(valor) for valor in re.findall(r"(?:19|20)\d{2}", nombre_hoja)}
    return tuple(sorted(anios))


def calcular_sha256(ruta: Path) -> str:
    """Calcula la identidad de la fuente por bloques para no cargarla en memoria."""

    digest = hashlib.sha256()
    with ruta.open("rb") as fuente:
        for bloque in iter(lambda: fuente.read(1024 * 1024), b""):
            digest.update(bloque)
    return digest.hexdigest()


def encontrar_encabezado(
    hoja: Any,
    columnas_requeridas: tuple[str, ...],
) -> tuple[int, tuple[str, ...]] | None:
    """Busca el encabezado y conserva la mejor coincidencia para reportar faltantes."""

    requeridas = {normalizar_etiqueta(columna) for columna in columnas_requeridas}
    limite = min(int(hoja.max_row) if hoja.max_row else 20, 20)
    mejor: tuple[int, int, tuple[str, ...]] | None = None
    for numero, fila in enumerate(
        hoja.iter_rows(min_row=1, max_row=limite, values_only=True),
        start=1,
    ):
        columnas = tuple(_texto(celda) for celda in fila)
        presentes = {normalizar_etiqueta(columna) for columna in columnas if columna}
        if requeridas <= presentes:
            return numero, columnas
        coincidencias = len(requeridas & presentes)
        if coincidencias and (mejor is None or coincidencias > mejor[0]):
            mejor = coincidencias, numero, columnas
    return (mejor[1], mejor[2]) if mejor else None


def _contar_filas(hoja: Any, fila_encabezado: int) -> int:
    """Cuenta filas con contenido y evita depender de espacios vacíos del XLSX."""

    cantidad = 0
    for fila in hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True):
        if any(_texto(celda) for celda in fila):
            cantidad += 1
    return cantidad


def _inspeccionar_hoja(
    hoja: Any,
    hallazgos: list[Hallazgo],
) -> HojaInspeccion:
    """Inspecciona una hoja y agrega errores específicos de columnas o encabezado."""

    nombre = str(hoja.title)
    rol = detectar_rol(nombre)
    anios = extraer_anios(nombre)
    if rol is None:
        hallazgos.append(
            Hallazgo(
                codigo="HOJA_NO_RECONOCIDA",
                severidad="warning",
                mensaje="La hoja no pertenece a un universo reconocido y no será procesada.",
                hoja=nombre,
            )
        )
        return HojaInspeccion(nombre, None, anios, None, 0, ())

    encontrado = encontrar_encabezado(hoja, COLUMNAS_OBLIGATORIAS[rol])
    if encontrado is None:
        hallazgos.append(
            Hallazgo(
                codigo="ENCABEZADO_NO_ENCONTRADO",
                severidad="error",
                mensaje="No se encontró una fila de encabezados compatible.",
                hoja=nombre,
                detalle="Columnas mínimas: " + ", ".join(COLUMNAS_OBLIGATORIAS[rol]),
            )
        )
        return HojaInspeccion(nombre, rol, anios, None, 0, ())

    fila_encabezado, columnas = encontrado
    presentes = {normalizar_etiqueta(columna) for columna in columnas if columna}
    faltantes = [
        columna
        for columna in COLUMNAS_OBLIGATORIAS[rol]
        if normalizar_etiqueta(columna) not in presentes
    ]
    if faltantes:
        hallazgos.append(
            Hallazgo(
                codigo="COLUMNAS_OBLIGATORIAS_AUSENTES",
                severidad="error",
                mensaje="Faltan columnas obligatorias para el universo identificado.",
                hoja=nombre,
                fila=fila_encabezado,
                detalle=", ".join(faltantes),
            )
        )

    return HojaInspeccion(
        nombre=nombre,
        rol=rol,
        anios=anios,
        encabezado_fila=fila_encabezado,
        filas_datos=_contar_filas(hoja, fila_encabezado),
        columnas=columnas,
    )


def validar_archivo(ruta: Path, nombre_archivo: str | None = None) -> ResultadoValidacionEntrada:
    """Valida un XLSX antes de cualquier limpieza semántica o llamada al LLM."""

    archivo = nombre_archivo or ruta.name
    hallazgos: list[Hallazgo] = []
    sha256 = calcular_sha256(ruta) if ruta.exists() else ""
    if not ruta.exists():
        hallazgos.append(
            Hallazgo(
                codigo="FUENTE_NO_ENCONTRADA",
                severidad="error",
                mensaje="No se encontró el archivo de entrada.",
                detalle=archivo,
            )
        )
        return ResultadoValidacionEntrada(archivo, sha256, False, (), tuple(hallazgos))

    if ruta.suffix.lower() != ".xlsx":
        hallazgos.append(
            Hallazgo(
                codigo="EXTENSION_INVALIDA",
                severidad="error",
                mensaje="La fuente debe ser un archivo XLSX.",
                detalle=ruta.suffix or "sin extensión",
            )
        )
        return ResultadoValidacionEntrada(archivo, sha256, False, (), tuple(hallazgos))

    try:
        libro = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as exc:
        hallazgos.append(
            Hallazgo(
                codigo="FUENTE_ILEGIBLE",
                severidad="error",
                mensaje="No se pudo abrir el XLSX.",
                detalle=f"{type(exc).__name__}: {str(exc)[:200]}",
            )
        )
        return ResultadoValidacionEntrada(archivo, sha256, False, (), tuple(hallazgos))

    hojas = tuple(_inspeccionar_hoja(libro[nombre], hallazgos) for nombre in libro.sheetnames)
    for rol in ROLES:
        candidatas = [hoja for hoja in hojas if hoja.rol == rol]
        if not candidatas:
            hallazgos.append(
                Hallazgo(
                    codigo="HOJA_REQUERIDA_AUSENTE",
                    severidad="error",
                    mensaje="No se encontró una hoja para el universo requerido.",
                    detalle=rol,
                )
            )
            continue

        # Las columnas adicionales están permitidas; cada hoja ya fue validada
        # contra el mismo conjunto mínimo de columnas obligatorias.

    try:
        libro.close()
    except Exception:
        pass

    valida = not any(hallazgo.severidad == "error" for hallazgo in hallazgos)
    return ResultadoValidacionEntrada(archivo, sha256, valida, hojas, tuple(hallazgos))
