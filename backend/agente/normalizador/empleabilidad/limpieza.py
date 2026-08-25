"""Limpieza estructural y staging reproducible de la fuente de Empleabilidad."""

from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Iterator
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from agente.normalizador.modelos import (
    ResultadoLimpieza,
    ResultadoValidacionEntrada,
    TipoHoja,
)

from .entrada import (
    COLUMNAS_OBLIGATORIAS,
    encontrar_encabezado,
    normalizar_etiqueta,
)

ROLES: tuple[TipoHoja, ...] = ("convenios", "informes", "publicaciones")


def _valor_limpio(valor: object) -> object:
    """Normaliza espacios, fechas y números sin aplicar interpretación semántica."""

    if valor is None:
        return None
    if isinstance(valor, (datetime, date, time)):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return int(valor)
    if isinstance(valor, str):
        return re.sub(r"\s+", " ", valor).strip()
    return valor


def _clave_campo(valor: object, indice: int) -> str:
    """Convierte un encabezado fuente a una clave JSON estable y legible."""

    clave = normalizar_etiqueta(valor).replace(" ", "_")
    return clave or f"columna_{indice}"


def _claves_unicas(columnas: tuple[str, ...]) -> tuple[str, ...]:
    """Evita que encabezados repetidos sobrescriban datos durante la limpieza."""

    cantidades: dict[str, int] = {}
    resultado: list[str] = []
    for indice, columna in enumerate(columnas, start=1):
        base = _clave_campo(columna, indice)
        cantidades[base] = cantidades.get(base, 0) + 1
        resultado.append(base if cantidades[base] == 1 else f"{base}_{cantidades[base]}")
    return tuple(resultado)


def _identificador_registro(
    rol: TipoHoja,
    nombre_hoja: str,
    fila: int,
    datos: dict[str, object],
) -> str:
    """Genera un ID reproducible que conserva unicidad aun con filas duplicadas."""

    payload = json.dumps(
        {"rol": rol, "hoja": nombre_hoja, "fila": fila, "datos": datos},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]
    return f"{rol}_{digest}"


def _iterar_hoja(hoja: Any, rol: TipoHoja) -> Iterator[tuple[int, dict[str, object]]]:
    """Itera filas no vacías conservando su número de origen."""

    encontrado = encontrar_encabezado(hoja, COLUMNAS_OBLIGATORIAS[rol])
    if encontrado is None:
        return
    fila_encabezado, columnas = encontrado
    claves = _claves_unicas(columnas)
    for numero, fila in enumerate(
        hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True),
        start=fila_encabezado + 1,
    ):
        datos = {
            clave: _valor_limpio(valor)
            for clave, valor in zip(claves, fila, strict=False)
        }
        if not any(valor is not None and valor != "" for valor in datos.values()):
            continue
        yield numero, datos


def _sha256(ruta: Path) -> str:
    """Calcula la huella del output sin cargarlo completo en memoria."""

    digest = hashlib.sha256()
    with ruta.open("rb") as archivo:
        for bloque in iter(lambda: archivo.read(1024 * 1024), b""):
            digest.update(bloque)
    return digest.hexdigest()


def limpiar_archivo(
    ruta_entrada: Path,
    directorio_ejecucion: Path,
    validacion: ResultadoValidacionEntrada,
) -> ResultadoLimpieza:
    """Genera un JSONL por universo, sin normalización semántica ni uso del LLM."""

    directorio_salida = directorio_ejecucion / "limpios"
    directorio_salida.mkdir(parents=True, exist_ok=True)
    registros: dict[str, int] = {}
    outputs: list[dict[str, object]] = []

    libro = load_workbook(ruta_entrada, read_only=True, data_only=True)
    try:
        for rol in ROLES:
            hojas = [hoja for hoja in validacion.hojas if hoja.rol == rol]
            destino = directorio_salida / f"{rol}.jsonl"
            cantidad = 0
            with destino.open("w", encoding="utf-8", newline="\n") as archivo:
                for inspeccion in hojas:
                    hoja = libro[inspeccion.nombre]
                    for numero, datos in _iterar_hoja(hoja, rol):
                        registro = {
                            "id_registro": _identificador_registro(
                                rol,
                                inspeccion.nombre,
                                numero,
                                datos,
                            ),
                            "universo": rol,
                            "origen": {"hoja": inspeccion.nombre, "fila": numero},
                            "datos": datos,
                        }
                        archivo.write(
                            json.dumps(registro, ensure_ascii=False, separators=(",", ":"))
                            + "\n"
                        )
                        cantidad += 1
            registros[rol] = cantidad
            outputs.append(
                {
                    "tipo": "staging_jsonl",
                    "universo": rol,
                    "archivo": f"limpios/{rol}.jsonl",
                    "registros": cantidad,
                    "bytes": destino.stat().st_size,
                    "sha256": _sha256(destino),
                }
            )
    finally:
        try:
            libro.close()
        except Exception:
            pass

    return ResultadoLimpieza(registros, tuple(outputs), ())
