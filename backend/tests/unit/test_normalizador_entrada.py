"""Pruebas del contrato estructural de entrada de Empleabilidad."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from agente.normalizador.empleabilidad.entrada import validar_archivo

CONVENIOS = ["RUC", "Empresa", "Facultad", "Cód_carrera", "Carrera", "Ciclo_convenio"]
INFORMES = ["Año", "Ciclo", "Facultad", "Cód_carrera", "Carrera"]
PUBLICACIONES = [
    "Año de la vacante",
    "RUC",
    "Razón Social",
    "Identificación",
    "Tipo de puesto",
    "Fecha de publicación",
    "Fecha de finalización",
    "Creado_Empleo",
    "Área",
    "Área específica",
    "Cargo",
    "Cargo específico",
    "Posición a publicar",
    "Carrera resumen",
    "Funciones",
]


def crear_fuente(ruta: Path, incluir_publicaciones: bool = True) -> None:
    """Crea una fuente pequeña con títulos y años variables."""

    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    hoja.title = "Convenios 2024 - 2026"
    hoja.append(["Fuente CIAR"])
    hoja.append(CONVENIOS)
    hoja.append(["20107798049", "Empresa de prueba", "Facultad", "01", "Administración", "6"])

    informes = libro.create_sheet("Informes 2021 - 2023")
    informes.append(INFORMES)
    informes.append(["2023", "6", "Facultad", "01", "Administración"])

    if incluir_publicaciones:
        publicaciones = libro.create_sheet("Publicaciones 2026")
        publicaciones.append(PUBLICACIONES)
        publicaciones.append(["2026"] + ["dato"] * (len(PUBLICACIONES) - 1))

    libro.save(ruta)


def test_acepta_universos_con_anios_variables(tmp_path: Path) -> None:
    """Los años del nombre de hoja son metadatos, no una restricción fija."""

    ruta = tmp_path / "fuente.xlsx"
    crear_fuente(ruta)

    resultado = validar_archivo(ruta)

    assert resultado.valida is True
    roles = {hoja.rol for hoja in resultado.hojas}
    assert roles == {"convenios", "informes", "publicaciones"}
    convenios = next(hoja for hoja in resultado.hojas if hoja.rol == "convenios")
    assert convenios.anios == (2024, 2026)
    assert convenios.encabezado_fila == 2


def test_reporta_hoja_obligatoria_ausente(tmp_path: Path) -> None:
    """La ausencia de un universo impide iniciar la normalización."""

    ruta = tmp_path / "fuente_incompleta.xlsx"
    crear_fuente(ruta, incluir_publicaciones=False)

    resultado = validar_archivo(ruta)

    assert resultado.valida is False
    assert any(
        hallazgo.codigo == "HOJA_REQUERIDA_AUSENTE"
        and hallazgo.detalle == "publicaciones"
        for hallazgo in resultado.hallazgos
    )


def test_reporta_columnas_faltantes(tmp_path: Path) -> None:
    """Una hoja reconocida pero incompleta produce un error específico."""

    ruta = tmp_path / "fuente_invalida.xlsx"
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    hoja.title = "Convenios 2030"
    hoja.append(["RUC", "Empresa"])
    libro.create_sheet("Informes 2030").append(INFORMES)
    libro.create_sheet("Publicaciones 2030").append(PUBLICACIONES)
    libro.save(ruta)

    resultado = validar_archivo(ruta)

    assert resultado.valida is False
    assert any(
        hallazgo.codigo == "COLUMNAS_OBLIGATORIAS_AUSENTES"
        and hallazgo.hoja == "Convenios 2030"
        for hallazgo in resultado.hallazgos
    )
